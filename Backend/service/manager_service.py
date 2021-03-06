import os
import random
import base64
from datetime import datetime
from flask import jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles.borders import Border, Side

from endpoints import Config
from .mail_service import Mailer
from dao import UserDao, PetitionDao, ManagerDao

class ManagerService:
    def __init__(self, user_dao:UserDao, petition_dao:PetitionDao, manager_dao:ManagerDao, mailer:Mailer, config:Config):
        self.user_dao = user_dao
        self.petition_dao = petition_dao
        self.manager_dao = manager_dao
        self.mailer= mailer
        self.config = config

    def get_user_count(self):
        return self.manager_dao.get_user_count()

    def get_petition_status(self, petition_id:int):
        title, status, supports = self.petition_dao.get_petition_status(petition_id)

        if status != None:
            return jsonify({
                "title":title,
                "status":status,
                "supports":supports,
                "msg":"success!"
            })
        else:
            return "청원번호 조회에 실패했습니다.", 400

    def get_report_service(self):
        return jsonify({ "reports":self.manager_dao.get_reports() })

    def deactivate_petition_service(self, pid:int):
        self.petition_dao.deactivate_petition(pid)
        return "성공!", 200

    def delete_user_service(self, uid:int):
        self.user_dao.update_withdraw(uid)
        return "성공!", 200

    def get_authcode_count(self):
        data = self.manager_dao.get_authcode_count()

        if "manager" not in data.keys():
            data["manager"] = 0
        if "general" not in data.keys():
            data["general"] = 0

        return jsonify(data)

    def generate_authcode_service(self, grade:int, count:int, priv:int, life:int):
        charset = "AB1CD3E2FG4HI5JK6LMN7OP8QR9STU0WXYZ"
        old_authcodes = self.user_dao.get_all_authcodes()
        authcodes = []

        grade = int(grade)
        count = int(count)
        priv = int(priv)
        life = int(life)

        if grade > 0 and grade < 4:
            count *= 30
        elif grade != 0:
            return "잘못된 접근입니다.", 400

        while len(authcodes) < count:
            authcode = ""
            for i in range(6):
                authcode += random.choice(charset)

            if authcode not in authcodes and authcode not in old_authcodes:
                authcodes.append(authcode)

        self.manager_dao.insert_authcode(grade, tuple(authcodes), priv, life)

        wb = Workbook()
        ws = wb.active
        border = Border(
            left = Side(style="thin"),
            right = Side(style="thin"),
            top = Side(style="thin"),
            bottom = Side(style="thin")
        )

        for i in range(len(authcodes)):
            rowidx = str((i % 30) + 3)

            if rowidx == "3":
                ws["B2"] = "학년"
                ws["C2"] = "인증번호"
                ws["D2"] = "권한"

                ws["B2"].border = border
                ws["C2"].border = border
                ws["D2"].border = border
                if not i < 30:
                    ws = wb.create_sheet()

            ws["B" + rowidx] = grade
            ws["C" + rowidx] = authcodes[i]
            ws["D" + rowidx] = "일반" if priv == 1 else "관리자"

            ws["B" + rowidx].border = border
            ws["C" + rowidx].border = border
            ws["D" + rowidx].border = border

        fname = f"{datetime.now().timestamp()}.xlsx"

        wb.save(fname)

        fp = open(fname, "rb")
        binary = base64.b64encode(fp.read()).decode("UTF-8")
        fp.close()

        os.remove(fname)

        return jsonify({
            "file":binary
        })

    def truncate_authcodes_service(self):
        self.manager_dao.truncate_authcodes()

    def write_notice_service(self, uid:int, title:str, content:str):
        if title and content:
            user_emails = self.user_dao.get_all_email()
            self.manager_dao.insert_notice(uid, title, content.replace("\n", "<br>"))

            if user_emails:
                mail_title = "[청원 시스템 공지] " + title
                self.mailer.send(mail_title, content, user_emails)
            return "성공!", 200
        else:
            return "빈칸을 모두 채워주세요.", 400

    def get_notice_metadata_service(self):
        return jsonify({ "notices":self.manager_dao.get_notice_metadata() })

    def get_notice(self, nid:int):
        data = self.manager_dao.get_notice(nid)

        if data == None:
            return "존재하지 않는 공지입니다.", 403
        else:
            return jsonify(data)

    def get_pass_line(self):
        return (self.manager_dao.get_user_count() *  self.config.pass_ratio) // 100